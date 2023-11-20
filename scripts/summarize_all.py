'''
Script to consolidate json results from log summarizer output to a single Excel file
'''
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill
import os
import sys
import glob
import json
import logging
import pandas as pd
import argparse

logging.basicConfig(stream=sys.stdout, level=logging.INFO, format='%(message)s')

class SheetTracker:
    CHAR_WIDTH = 2
    # SheetTracker for a measure stores dataset:column and experiment:row info
    def __init__(self, wb: Workbook, measure: str):
        self.sheet = wb.create_sheet(measure)
        # First row/column is for labels
        self.dataset_column_map = {'header': 1}
        self.exp_row_map = {'header': 1}
        self.exp_label_size = 10

    def add_entry(self, exp, dataset, value):
        if exp not in self.exp_row_map:
            self.exp_row_map[exp] = len(self.exp_row_map) + 1
            self.sheet.cell(self.exp_row_map[exp], 1, exp)
            self.exp_label_size = max(self.exp_label_size, len(dataset))
            self.sheet.column_dimensions['A'].width = self.exp_label_size * self.CHAR_WIDTH
        if dataset not in self.dataset_column_map:
            new_column_number = len(self.dataset_column_map) + 1
            self.dataset_column_map[dataset] = new_column_number
            self.sheet.cell(1, new_column_number, dataset)
            self.sheet.column_dimensions[get_column_letter(new_column_number)].width = 5 * self.CHAR_WIDTH

        self.sheet.cell(self.exp_row_map[exp], self.dataset_column_map[dataset], value)


    def _add_average_column(self):
        # Add column for average value
        num_rows = len(self.exp_row_map)
        # Setup new AVG column
        self.add_entry('header', 'AVG Val', 'AVG Val')
        avg_column_number = self.dataset_column_map['AVG Val']
        last_column_letter = get_column_letter(avg_column_number - 1)
        for row_ind in range(2, num_rows+1):
            data_range = self.sheet['B'+str(row_ind):last_column_letter+str(row_ind)]
            vals = [cell.value for row in data_range for cell in row if cell.value is not None]
            average_val = sum(vals) / len(vals)
            self.sheet.cell(row_ind, avg_column_number, float(f'{average_val:.1f}'))


    def _get_graded_color(self, index: int, start_color: [int], color_step: [float]):
        # Calculate and return cell fill based on start color, color step, and rank index
        fill_color = Color(rgb=f'{int(start_color[0] + index*color_step[0]):02X}'
            f'{int(start_color[1] + index*color_step[1]):02X}'
            f'{int(start_color[2] + index*color_step[2]):02X}'
        )
        return PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')


    def _add_rank_info_to_cells(self):
        # Add rank of the cell in brackets column-wise and a column for average rank
        # Highest value in a column is boldened
        num_rows = len(self.exp_row_map)
        row_ranks_map = {row: [] for row in range(2, num_rows+1)}
        start_color = (0, 255, 0)
        end_color = (255, 255, 0)
        color_step = [(end - start) / num_rows for start, end in zip(start_color, end_color)]

        for column_ind in range(2, len(self.dataset_column_map)+1):
            column_letter = get_column_letter(column_ind)
            data_range = self.sheet[column_letter+'2':column_letter+str(num_rows)]
            # data_range is 2D, make it a list
            cells = [row[0] for row in data_range if row[0].value is not None]
            sorted_cells = sorted(cells, key=lambda cell: cell.value, reverse=True)

            sorted_cells[0].font = openpyxl.styles.Font(bold=True)
            for ind, cell in enumerate(sorted_cells):
                cell.value = f'{cell.value:2.2f} ({ind+1})'
                row_ranks_map[cell.row].append(ind+1)
                cell.fill = self._get_graded_color(ind, start_color, color_step)

        # Setup avg rank column
        self.add_entry('header', 'AVG Rank', 'AVG Rank')
        avg_rank_column = self.dataset_column_map['AVG Rank']
        row_avg_ranks = [(row_ind, sum(ranks)/len(ranks)) for row_ind, ranks in row_ranks_map.items()]
        sorted_row_avg_ranks = sorted(row_avg_ranks, key=lambda val: val[1])

        top_rank_cell = self.sheet.cell(sorted_row_avg_ranks[0][0], avg_rank_column)
        top_rank_cell.font = openpyxl.styles.Font(bold=True)
        for ind, (row_ind, avg_rank) in enumerate(sorted_row_avg_ranks):
            cell = self.sheet.cell(row_ind, avg_rank_column, f'{avg_rank:.2f} ({ind+1})')
            cell.fill = self._get_graded_color(ind, start_color, color_step)


    def finalize(self):
        self._add_average_column()
        self._add_rank_info_to_cells()


def process_summary_data(wb: Workbook, measure_sheet_map: dict, data: dict):
    for measure in data['scores']:
        if measure not in measure_sheet_map:
            measure_sheet_map[measure] = SheetTracker(wb, measure)
        score, variance = data['scores'][measure]
        experiment = data['experiment']
        dataset = data['dataset']
        # measure_sheet_map[measure].add_entry(experiment, dataset, f'{score:.2f}\u00B1{variance:.2f}')
        measure_sheet_map[measure].add_entry(experiment, dataset, float(f'{score:.3f}')*100)


def add_template_data(measure_sheet_map: dict, template_path: str):
    df = pd.read_csv(template_path)
    datasets = [col for col in df if col not in ['measure', 'experiment']]
    for _, row in df.iterrows():
        sheet_tracker = measure_sheet_map[row['measure']]
        experiment = row['experiment']
        for dataset in datasets:
            sheet_tracker.add_entry(experiment, dataset, row[dataset])


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input_file', help='Path to input json file', type=str, default=os.path.join('results', 'summary.json'))
    parser.add_argument('-o', '--output_file', help='Path to output excel file', type=str, default=os.path.join('results', 'summary.xlsx'))
    parser.add_argument('-t', '--template_file', action='append', help='Add csv file with additional info to add on')
    parser.add_argument('-v', '--verbose', help='Increase output verbosity', action='store_true')
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    wb = Workbook()
    # Delete default sheet
    del wb['Sheet']
    measure_sheet_map = dict()

    with open(args.input_file) as ff:
        logging.info(f'Processing data from file {args.input_file}')
        data = json.load(ff)
        for entry in data:
            process_summary_data(wb, measure_sheet_map, entry)

    for template_file in args.template_file:
        logging.info(f'Adding additional info from {template_file}')
        add_template_data(measure_sheet_map, 'smd_template.csv')

    for measure, sheet_tracker in measure_sheet_map.items():
        sheet_tracker.finalize()

    logging.info(f'Writing result to {args.output_file}')
    wb.save(args.output_file)

